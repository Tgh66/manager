import os
import sqlite3
import hashlib
from datetime import datetime
import io
import base64
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from flask import g
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image as PILImage

#
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # 用于会话管理的密钥
DATABASE = 'users.db'
UPLOAD_FOLDER = 'uploads'
EXCEL_FOLDER = 'excel_files'

# 确保上传和Excel文件夹存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)


# 数据库连接函数
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db


# 初始化数据库
def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        # 创建用户表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        # 添加管理员表
        cursor.execute('''
                CREATE TABLE IF NOT EXISTS admins (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
                ''')

        # 检查是否有管理员账号，如果没有则创建默认管理员
        cursor.execute('SELECT * FROM admins WHERE username = ?', ('admin',))
        if not cursor.fetchone():
            default_pwd_hash = hash_password('123')  # 默认密码123
            cursor.execute(
                'INSERT INTO admins (username, password_hash) VALUES (?, ?)',
                ('admin', default_pwd_hash)
            )
        db.commit()


# 关闭数据库连接
@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


# 哈希密码
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# 检查密码
def check_password(password, hashed):
    return hash_password(password) == hashed


# 注册路由
@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    room = data.get('room')
    password = data.get('password')

    if not room or not password:
        return jsonify({'success': False, 'message': '房间号和密码不能为空'})

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute('INSERT INTO users (room_number, password_hash) VALUES (?, ?)',
                       (room, hash_password(password)))
        db.commit()
        return jsonify({'success': True, 'message': '注册成功'})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': '该房间号已注册'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'注册失败: {str(e)}'})


# 登录路由
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    room = data.get('room')
    password = data.get('password')

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM users WHERE room_number = ?', (room,))
    user = cursor.fetchone()

    if user and check_password(password, user['password_hash']):
        session['room'] = room
        return jsonify({'success': True, 'message': '登录成功'})
    else:
        return jsonify({'success': False, 'message': '房间号或密码错误'})


# 退出登录
@app.route('/logout')
def logout():
    session.pop('room', None)
    return jsonify({'success': True, 'message': '已退出登录'})


# 获取当前登录用户
@app.route('/get_current_user')
def get_current_user():
    if 'room' in session:
        return jsonify({'success': True, 'room': session['room']})
    else:
        return jsonify({'success': False, 'message': '未登录'})


# 首页路由（登录页面）
@app.route('/')
def index():
    if 'room' in session:
        return redirect(url_for('user'))
    return render_template('login.html')


# 用户页面（表单填写）
@app.route('/user')
def user():
    if 'room' not in session:
        return redirect(url_for('index'))
    return render_template('user.html')


# 下载详细字段文档
@app.route('/download_fields_doc')
def download_fields_doc():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    # 假设文档与app.py在同一目录
    doc_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '详细字段.docx')

    if not os.path.exists(doc_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    try:
        return send_file(
            doc_path,
            as_attachment=True,
            download_name='详细字段.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'})


# 保存上传的图片
def save_image(file):
    if file and file.filename != '':
        # 生成唯一文件名
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
        filename = f"{timestamp}_{file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)

        # 保存文件
        file.save(filepath)
        return filepath
    return None


# 将图片插入到Excel
def insert_image_to_excel(ws, image_path, row, col, max_width=300, max_height=200):
    if not image_path or not os.path.exists(image_path):
        return

    try:
        # 打开并调整图片大小
        img = PILImage.open(image_path)
        img.thumbnail((max_width, max_height))

        # 保存调整后的图片到临时内存
        temp_img = io.BytesIO()
        img.save(temp_img, format='PNG')
        temp_img.seek(0)

        # 插入到Excel
        excel_img = Image(temp_img)
        ws.add_image(excel_img, f"{get_column_letter(col)}{row}")

        # 调整行高和列宽以适应图片
        ws.row_dimensions[row].height = img.height * 0.75  # 行高大约是像素的0.75倍
        ws.column_dimensions[get_column_letter(col)].width = img.width * 0.14  # 列宽大约是像素的0.14倍
    except Exception as e:
        print(f"插入图片出错: {str(e)}")


# 获取最后一次数据回填
@app.route('/get_last_submission')
def get_last_submission():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    room = session['room']
    excel_filename = f"{room}.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    if not os.path.exists(excel_path):
        return jsonify({'success': False, 'message': '没有历史数据'})

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        # 获取最新的工作表（按创建时间排序）
        # 排除默认工作表（如果存在）
        sheets = [sheet for sheet in wb.sheetnames if sheet != 'Sheet']
        if not sheets:
            return jsonify({'success': False, 'message': '没有历史数据'})

        # 按工作表名（包含时间戳）排序，取最后一个
        sheets.sort()
        last_sheet_name = sheets[-1]
        ws = wb[last_sheet_name]

        # 解析表单数据
        form_data = {}
        current_row = 3  # 从项目负责人信息开始

        # 解析项目负责人信息
        leader_fields = [
            ("projectLeaderName", "项目负责人姓名"),
            ("projectLeaderCollege", "项目负责人学院"),
            ("projectLeaderGrade", "项目负责人年级"),
            ("projectLeaderGender", "项目负责人性别"),
            ("projectLeaderPhone", "项目负责人联系电话"),
            ("projectType", "项目类型")
        ]

        for field_name, field_label in leader_fields:
            # 查找标签所在行
            while current_row <= ws.max_row:
                cell_value = ws.cell(row=current_row, column=1).value
                if cell_value == field_label:
                    form_data[field_name] = ws.cell(row=current_row, column=2).value
                    current_row += 1
                    break
                current_row += 1

        # 如果是在孵企业，解析企业信息
        project_type = form_data.get('projectType')
        if project_type == '在孵企业':
            form_data['projectType'] = '1'  # 转换为表单值
            current_row += 1  # 跳过"企业信息"标题行

            enterprise_fields = [
                ("enterpriseAccount", "在孵企业帐号(18位统一社会信用代码)"),
                ("enterpriseName", "企业名称"),
                ("establishmentDate", "企业成立时间"),
                ("registeredCapital", "企业成立时注册资本(千元)"),
                ("incubationStartDate", "企业入驻时间"),
                ("areaOccupied", "占用孵化器场地面积(平方米)"),
                ("registrationType", "企业登记注册类型"),
                ("techField", "企业所属技术领域"),
                ("coreTechField1", "企业核心技术所属领域 - 大类"),
                ("coreTechField2", "企业核心技术所属领域 - 中类"),
                ("coreTechField3", "企业核心技术所属领域 - 小类"),
                ("industryCategory1", "行业类别 - 大类"),
                ("industryCategory2", "行业类别 - 中类"),
                ("industryCategory3", "行业类别 - 小类"),
                ("industryCategory4", "行业类别 - 细类"),
                ("taxpayerType", "企业纳税人类型"),
                ("totalRevenue", "在孵企业总收入(千元)"),
                ("netProfit", "在孵企业净利润(千元)"),
                ("exportAmount", "在孵企业出口总额(千元)"),
                ("rdExpenditure", "研究与试验发展经费(千元)"),
                ("taxPayment", "实际上缴税费(千元)")
            ]

            for field_name, field_label in enterprise_fields:
                while current_row <= ws.max_row:
                    cell_value = ws.cell(row=current_row, column=1).value
                    if cell_value == field_label:
                        form_data[field_name] = ws.cell(row=current_row, column=2).value
                        current_row += 1
                        break
                    current_row += 1

        # 解析项目成员信息
        while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value != "项目成员信息":
            current_row += 1
        current_row += 1  # 跳过标题行
        current_row += 1  # 跳过头行
        # 跳过成员数据行，直到下一个标题
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if cell_value in ["知识产权信息", "赛事获奖信息"]:
                break
            current_row += 1

        # 解析赛事获奖信息
        form_data["awards"] = []
        while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value != "赛事获奖信息":
            current_row += 1
        if current_row <= ws.max_row and ws.cell(row=current_row, column=1).value == "赛事获奖信息":
            current_row += 1  # 跳过标题行
            current_row += 1  # 跳过头行

            while current_row <= ws.max_row:
                cell_value = ws.cell(row=current_row, column=1).value
                # 检查是否到达下一个信息板块
                if cell_value in ["知识产权信息", "企业资质信息", "投融资信息"]:
                    break
                if cell_value and str(cell_value).isdigit():
                    award = {
                        "competition": ws.cell(row=current_row, column=2).value,
                        "prize": ws.cell(row=current_row, column=3).value
                    }
                    form_data["awards"].append(award)
                current_row += 1

        # 解析知识产权信息
        while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value != "知识产权信息":
            current_row += 1
        current_row += 1  # 跳过标题行

        ip_fields = [
            ("ipApplications", "当年知识产权申请数(件)"),
            ("ipAuthorizations", "当年知识产权授权数(件)"),
            ("inventionPatents", "其中：发明专利(件)"),
            ("softwareCopyrights", "软件著作权(件)"),
            ("techContracts", "技术合同成交数量(项)"),
            ("techContractAmount", "技术合同成交额(千元)"),
            ("nationalProjects", "当年承担国家级科技计划项目数(项)")
        ]

        for field_name, field_label in ip_fields:
            while current_row <= ws.max_row:
                cell_value = ws.cell(row=current_row, column=1).value
                if cell_value == field_label:
                    form_data[field_name] = ws.cell(row=current_row, column=2).value
                    current_row += 1
                    break
                current_row += 1

        # 解析企业资质信息
        while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value != "企业资质信息":
            current_row += 1
        current_row += 1  # 跳过标题行

        qualification_fields = [
            ("isHighTechEnterprise", "是否高新技术企业"),
            ("highTechCertificateNo", "高新技术企业证书编号"),
            ("isTechSme", "是否是科技型中小企业"),
            ("techSmeCode", "科技型中小企业登记编码"),
            ("isInnovativeSme", "是否创新型中小企业"),
            ("isSpecializedSme", "是否专精特新中小企业"),
            ("isGiantSme", "是否专精特新“小巨人”企业")
        ]

        for field_name, field_label in qualification_fields:
            while current_row <= ws.max_row:
                cell_value = ws.cell(row=current_row, column=1).value
                if cell_value == field_label:
                    val = ws.cell(row=current_row, column=2).value
                    # 转换为表单值
                    if val == "是":
                        form_data[field_name] = "yes"
                    elif val == "否":
                        form_data[field_name] = "no"
                    else:
                        form_data[field_name] = val
                    current_row += 1
                    break
                current_row += 1

        # 解析投融资信息
        while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value != "投融资信息":
            current_row += 1
        current_row += 1  # 跳过标题行

        finance_fields = [
            ("financingAmount", "获得投融资金额(千元)"),
            ("incubatorFundAmount", "其中：获得孵化器孵化基金投资额(千元)"),
            ("bankLoanAmount", "其中：获银行贷款额(千元)")
        ]

        for field_name, field_label in finance_fields:
            while current_row <= ws.max_row:
                cell_value = ws.cell(row=current_row, column=1).value
                if cell_value == field_label:
                    form_data[field_name] = ws.cell(row=current_row, column=2).value
                    current_row += 1
                    break
                current_row += 1

        wb.close()
        return jsonify({'success': True, 'data': form_data})

    except Exception as e:
        print(f"获取最后一次提交数据出错: {str(e)}")
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'})


# 提交表单处理
@app.route('/submit_form', methods=['POST'])
def submit_form():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    room = session['room']
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    excel_filename = f"{room}.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    try:
        # 保存上传的图片
        business_license_path = save_image(request.files.get('businessLicense'))
        invention_patent_path = save_image(request.files.get('inventionPatentCertificate'))
        software_copyright_path = save_image(request.files.get('softwareCopyrightCertificate'))

        # 保存赛事获奖证明图片
        award_certificate_paths = []
        award_certificates = request.files.getlist('award_certificate[]')
        for cert in award_certificates:
            path = save_image(cert)
            award_certificate_paths.append(path)

        # 检查Excel文件是否存在，不存在则创建
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        # 创建新的工作表，以时间戳命名
        sheet_name = timestamp.replace(':', '-')  # 替换冒号，Excel不允许工作表名包含冒号
        if len(sheet_name) > 31:  # Excel工作表名最大长度为31
            sheet_name = sheet_name[:31]

        # 如果工作表名已存在，添加后缀
        counter = 1
        original_sheet_name = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        # 记录当前行号
        current_row = 1

        # 添加提交时间
        ws.cell(row=current_row, column=1, value="提交时间")
        ws.cell(row=current_row, column=2, value=timestamp)
        current_row += 2

        # 项目负责人信息 - 修复字体设置方式
        header_cell = ws.cell(row=current_row, column=1, value="项目负责人信息")
        header_cell.font = get_header_font()  # 先创建单元格，再设置字体
        current_row += 1

        fields = [
            ("projectLeaderName", "项目负责人姓名"),
            ("projectLeaderCollege", "项目负责人学院"),
            ("projectLeaderGrade", "项目负责人年级"),
            ("projectLeaderGender", "项目负责人性别", {"male": "男", "female": "女"}),
            ("projectLeaderPhone", "项目负责人联系电话"),
            ("projectType", "项目类型", {"1": "在孵企业", "2": "创业团队"})
        ]

        current_row = add_fields_to_excel(ws, current_row, fields, request.form)
        current_row += 1

        # 如果是在孵企业，添加企业信息
        project_type = request.form.get('projectType')
        if project_type == '1':
            header_cell = ws.cell(row=current_row, column=1, value="企业信息")
            header_cell.font = get_header_font()
            current_row += 1

            enterprise_fields = [
                ("enterpriseAccount", "在孵企业帐号(18位统一社会信用代码)"),
                ("enterpriseName", "企业名称"),
                ("establishmentDate", "企业成立时间"),
                ("registeredCapital", "企业成立时注册资本(千元)"),
                ("incubationStartDate", "企业入驻时间"),
                ("areaOccupied", "占用孵化器场地面积(平方米)"),
                ("registrationType", "企业登记注册类型", get_registration_type_map()),
                ("techField", "企业所属技术领域"),
                ("coreTechField1", "企业核心技术所属领域 - 大类"),
                ("coreTechField2", "企业核心技术所属领域 - 中类"),
                ("coreTechField3", "企业核心技术所属领域 - 小类"),
                ("industryCategory1", "行业类别 - 大类"),
                ("industryCategory2", "行业类别 - 中类"),
                ("industryCategory3", "行业类别 - 小类"),
                ("industryCategory4", "行业类别 - 细类"),
                ("taxpayerType", "企业纳税人类型", {"general": "一般纳税人", "small": "小规模纳税人"}),
                ("totalRevenue", "在孵企业总收入(千元)"),
                ("netProfit", "在孵企业净利润(千元)"),
                ("exportAmount", "在孵企业出口总额(千元)"),
                ("rdExpenditure", "研究与试验发展经费(千元)"),
                ("taxPayment", "实际上缴税费(千元)")
            ]

            current_row = add_fields_to_excel(ws, current_row, enterprise_fields, request.form)
            current_row += 1

            # 插入营业执照图片
            if business_license_path:
                ws.cell(row=current_row, column=1, value="营业执照照片")
                insert_image_to_excel(ws, business_license_path, current_row, 2)
                current_row += 5  # 留出空间给图片

        # 项目成员信息
        header_cell = ws.cell(row=current_row, column=1, value="项目成员信息")
        header_cell.font = get_header_font()
        current_row += 1

        # 获取所有成员信息
        member_names = request.form.getlist('member_name[]')
        member_genders = request.form.getlist('member_gender[]')
        member_is_students = request.form.getlist('member_isStudent[]')
        member_colleges = request.form.getlist('member_college[]')
        member_grades = request.form.getlist('member_grade[]')
        member_levels = request.form.getlist('member_level[]')
        member_phones = request.form.getlist('member_phone[]')
        member_is_overseas = request.form.getlist('member_isOverseas[]')

        # 成员表头
        member_headers = ["序号", "姓名", "性别", "是否在校生", "学院", "年级", "层次", "联系电话", "是否留学人员"]
        for col, header in enumerate(member_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1

        # 成员数据
        gender_map = {"male": "男", "female": "女"}
        yes_no_map = {"yes": "是", "no": "否"}
        level_map = {"undergraduate": "本科", "junior": "专科"}

        for i in range(len(member_names)):
            ws.cell(row=current_row, column=1, value=i + 1)
            ws.cell(row=current_row, column=2, value=member_names[i])
            ws.cell(row=current_row, column=3, value=gender_map.get(member_genders[i], ""))
            ws.cell(row=current_row, column=4, value=yes_no_map.get(member_is_students[i], ""))
            ws.cell(row=current_row, column=5, value=member_colleges[i])
            ws.cell(row=current_row, column=6, value=member_grades[i])
            ws.cell(row=current_row, column=7, value=level_map.get(member_levels[i], ""))
            ws.cell(row=current_row, column=8, value=member_phones[i])
            ws.cell(row=current_row, column=9, value=yes_no_map.get(member_is_overseas[i], ""))
            current_row += 1

        current_row += 1

        # 赛事获奖信息
        header_cell = ws.cell(row=current_row, column=1, value="赛事获奖信息")
        header_cell.font = get_header_font()
        current_row += 1

        # 获取所有赛事获奖信息
        award_competitions = request.form.getlist('award_competition[]')
        award_prizes = request.form.getlist('award_prize[]')

        # 获奖记录表头
        award_headers = ["序号", "赛事完整名称", "所获奖项", "图片证明"]
        for col, header in enumerate(award_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1

        # 处理获奖记录和图片
        for i in range(len(award_competitions)):
            ws.cell(row=current_row, column=1, value=i + 1)
            ws.cell(row=current_row, column=2, value=award_competitions[i])
            ws.cell(row=current_row, column=3, value=award_prizes[i])

            # 记录图片状态
            if i < len(award_certificate_paths) and award_certificate_paths[i]:
                ws.cell(row=current_row, column=4, value="有图片")
            else:
                ws.cell(row=current_row, column=4, value="无")

            current_row += 1

        current_row += 1

        # 插入获奖证明图片
        for i, img_path in enumerate(award_certificate_paths):
            if img_path:
                ws.cell(row=current_row, column=1, value=f"获奖记录 {i + 1} 证明图片")
                insert_image_to_excel(ws, img_path, current_row, 2)
                current_row += 5  # 留出空间给图片

        # 知识产权信息
        header_cell = ws.cell(row=current_row, column=1, value="知识产权信息")
        header_cell.font = get_header_font()
        current_row += 1

        ip_fields = [
            ("ipApplications", "当年知识产权申请数(件)"),
            ("ipAuthorizations", "当年知识产权授权数(件)"),
            ("inventionPatents", "其中：发明专利(件)"),
            ("softwareCopyrights", "软件著作权(件)"),
            ("techContracts", "技术合同成交数量(项)"),
            ("techContractAmount", "技术合同成交额(千元)"),
            ("nationalProjects", "当年承担国家级科技计划项目数(项)")
        ]

        current_row = add_fields_to_excel(ws, current_row, ip_fields, request.form)
        current_row += 1

        # 插入发明专利证书图片
        if invention_patent_path and int(request.form.get('inventionPatents', 0)) > 0:
            ws.cell(row=current_row, column=1, value="发明专利证书")
            insert_image_to_excel(ws, invention_patent_path, current_row, 2)
            current_row += 5  # 留出空间给图片

        # 插入软件著作权证书图片
        if software_copyright_path and int(request.form.get('softwareCopyrights', 0)) > 0:
            ws.cell(row=current_row, column=1, value="软件著作权证书")
            insert_image_to_excel(ws, software_copyright_path, current_row, 2)
            current_row += 5  # 留出空间给图片

        # 企业资质信息
        header_cell = ws.cell(row=current_row, column=1, value="企业资质信息")
        header_cell.font = get_header_font()
        current_row += 1

        qualification_fields = [
            ("isHighTechEnterprise", "是否高新技术企业", yes_no_map),
            ("highTechCertificateNo", "高新技术企业证书编号"),
            ("isTechSme", "是否是科技型中小企业", yes_no_map),
            ("techSmeCode", "科技型中小企业登记编码"),
            ("isInnovativeSme", "是否创新型中小企业", yes_no_map),
            ("isSpecializedSme", "是否专精特新中小企业", yes_no_map),
            ("isGiantSme", "是否专精特新“小巨人”企业", yes_no_map)
        ]

        current_row = add_fields_to_excel(ws, current_row, qualification_fields, request.form)
        current_row += 1

        # 投融资信息
        header_cell = ws.cell(row=current_row, column=1, value="投融资信息")
        header_cell.font = get_header_font()
        current_row += 1

        finance_fields = [
            ("financingAmount", "获得投融资金额(千元)"),
            ("incubatorFundAmount", "其中：获得孵化器孵化基金投资额(千元)"),
            ("bankLoanAmount", "其中：获银行贷款额(千元)")
        ]

        current_row = add_fields_to_excel(ws, current_row, finance_fields, request.form)

        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        # 保存Excel文件
        wb.save(excel_path)

        # 清理临时图片文件
        all_paths = [business_license_path, invention_patent_path, software_copyright_path] + award_certificate_paths
        for path in all_paths:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except:
                    pass

        return jsonify({'success': True, 'message': '表单提交成功'})
    except Exception as e:
        print(f"提交表单出错: {str(e)}")
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'})


# 辅助函数：获取表头字体样式
def get_header_font():
    return Font(bold=True, size=12)


# 辅助函数：获取企业登记注册类型映射
def get_registration_type_map():
    return {
        "110": "110.国有",
        "120": "120.集体",
        "130": "130.股份合作",
        "141": "141.国有联营",
        "142": "142.集体联营",
        "143": "143.国有与集体联营",
        "149": "149.其他联营",
        "151": "151.国有独资公司",
        "159": "159.其他有限责任公司",
        "160": "160.股份有限公司",
        "171": "171.私营独资",
        "172": "172.私营合伙",
        "173": "173.私营有限责任",
        "174": "174.私营股份有限",
        "190": "190.其他",
        "210": "210.与港澳台商合资经营",
        "220": "220.与港澳台商合作经营",
        "230": "230.港澳台商独资",
        "240": "240.港澳台商投资股份有限公司",
        "290": "290.其他港澳台商投资",
        "310": "310.中外合资经营",
        "320": "320.中外合作经营",
        "330": "330.外资企业",
        "340": "340.外商投资股份有限公司",
        "390": "390.其他外商投资"
    }


# 辅助函数：添加字段到Excel
def add_fields_to_excel(ws, start_row, fields, form_data):
    current_row = start_row
    for field_info in fields:
        field_name = field_info[0]
        field_label = field_info[1]
        value_map = field_info[2] if len(field_info) > 2 else None

        value = form_data.get(field_name, "")
        if value_map and value in value_map:
            value = value_map[value]

        ws.cell(row=current_row, column=1, value=field_label)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    return current_row


# 获取历史记录
@app.route('/get_history')
def get_history():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    room = session['room']
    excel_filename = f"{room}.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    try:
        if not os.path.exists(excel_path):
            return jsonify({'success': True, 'records': []})

        wb = load_workbook(excel_path, read_only=True)
        # 排除默认的Sheet
        sheet_names = [name for name in wb.sheetnames if name != 'Sheet']
        # 转换回原始时间戳格式
        records = []
        for name in sheet_names:
            timestamp = name.replace('-', ':').rsplit('_', 1)[0]  # 还原冒号，去除可能的计数器
            records.append({'timestamp': timestamp})

        # 按时间戳排序（最新的在前）
        records.sort(key=lambda x: x['timestamp'], reverse=True)

        return jsonify({'success': True, 'records': records})
    except Exception as e:
        print(f"获取历史记录出错: {str(e)}")
        return jsonify({'success': False, 'message': f'获取历史记录失败: {str(e)}'})


# 下载Excel文件
@app.route('/download_excel')
def download_excel():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    room = session['room']
    timestamp = request.args.get('timestamp')
    excel_filename = f"{room}.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    if not os.path.exists(excel_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    try:
        # 创建一个临时Excel文件，只包含请求的工作表
        wb = load_workbook(excel_path)
        sheet_name = timestamp.replace(':', '-')  # 转换为工作表名格式

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            # 检查带有计数器的版本
            found = False
            counter = 1
            while not found and counter <= 100:  # 限制最大尝试次数
                temp_name = f"{sheet_name}_{counter}"
                if temp_name in wb.sheetnames:
                    sheet_name = temp_name
                    found = True
                counter += 1

            if not found:
                return jsonify({'success': False, 'message': '记录不存在'})

        # 删除其他工作表
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]

        # 保存到临时内存
        temp_file = io.BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)

        # 提供下载
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"{room}_{timestamp.replace(':', '-')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"下载Excel出错: {str(e)}")
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'})


# 获取记录详情
@app.route('/get_record')
def get_record():
    if 'room' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    return jsonify({'success': True, 'message': '记录详情'})


from flask import Blueprint, render_template, request, jsonify, session, send_file, redirect, url_for
import os
import io
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image

# 管理员蓝图
admin_bp = Blueprint('admin', __name__)


# 管理员登录页面
@admin_bp.route('/admin/login')
def admin_login():
    if 'admin_logged_in' in session:
        return redirect(url_for('admin.admin_panel'))
    return render_template('admin_login.html')


# 管理员登录处理
@admin_bp.route('/admin/login', methods=['POST'])
def admin_login_process():
    data = request.get_json()  # 修改为获取JSON数据，与前端保持一致
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'})

    # 从数据库验证管理员
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM admins WHERE username = ?', (username,))
    admin = cursor.fetchone()

    if admin and check_password(password, admin['password_hash']):
        session['admin_logged_in'] = True
        return jsonify({'success': True, 'message': '登录成功'})
    else:
        return jsonify({'success': False, 'message': '用户名或密码错误'})


# 管理员面板
@admin_bp.route('/admin/panel')
def admin_panel():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin.admin_login'))
    return render_template('admin_panel.html')


# 管理员登出
@admin_bp.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin.admin_login'))


# 获取所有房间的表单记录
@admin_bp.route('/admin/get_all_rooms')
def get_all_rooms():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    try:
        # 获取所有房间的Excel文件
        rooms = []
        if os.path.exists(EXCEL_FOLDER) and os.path.isdir(EXCEL_FOLDER):
            for filename in os.listdir(EXCEL_FOLDER):
                if filename.endswith('.xlsx') and filename != 'Sheet.xlsx':
                    room_number = filename[:-5]  # 去除.xlsx扩展名
                    excel_path = os.path.join(EXCEL_FOLDER, filename)

                    # 获取该房间的所有表单记录
                    wb = load_workbook(excel_path, read_only=True)
                    # 排除默认的Sheet
                    sheet_names = [name for name in wb.sheetnames if name != 'Sheet']

                    records = []
                    for name in sheet_names:
                        timestamp = name.replace('-', ':').rsplit('_', 1)[0]
                        records.append({
                            'timestamp': timestamp,
                            'sheet_name': name
                        })

                    # 按时间戳排序
                    records.sort(key=lambda x: x['timestamp'], reverse=True)

                    rooms.append({
                        'room_number': room_number,
                        'records': records
                    })

        # 按房间号从小到大排序
        rooms.sort(key=lambda x: int(x['room_number']))

        return jsonify({'success': True, 'rooms': rooms})
    except Exception as e:
        print(f"获取房间记录出错: {str(e)}")
        return jsonify({'success': False, 'message': f'获取记录失败: {str(e)}'})


# 管理员下载单个表单
@admin_bp.route('/admin/download_single')
def download_single():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    room = request.args.get('room')
    sheet_name = request.args.get('sheet_name')

    if not room or not sheet_name:
        return jsonify({'success': False, 'message': '参数缺失'})

    excel_filename = f"{room}.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    if not os.path.exists(excel_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    try:
        # 创建临时文件，只包含请求的工作表
        wb = load_workbook(excel_path)

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            return jsonify({'success': False, 'message': '记录不存在'})

        # 删除其他工作表
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]

        # 保存到临时内存
        temp_file = io.BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)

        # 提供下载
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"room_{room}_{sheet_name}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"下载表单出错: {str(e)}")
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'})


# 管理员批量下载表单
@admin_bp.route('/admin/download_batch', methods=['POST'])
def download_batch():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': '请先登录'})

    data = request.get_json()
    selected_records = data.get('records', [])

    if not selected_records:
        return jsonify({'success': False, 'message': '请选择要下载的记录'})

    try:
        # 创建一个新的Excel工作簿
        wb = Workbook()
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 遍历选中的记录，添加到新工作簿
        for record in selected_records:
            room = record.get('room')
            sheet_name = record.get('sheet_name')

            if not room or not sheet_name:
                continue

            excel_filename = f"{room}.xlsx"
            excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

            if not os.path.exists(excel_path):
                continue

            # 打开源工作簿（需启用read_only=False以支持图片操作）
            source_wb = load_workbook(excel_path, read_only=False, data_only=True)
            if sheet_name not in source_wb.sheetnames:
                source_wb.close()
                continue

            # 复制工作表到新工作簿
            source_ws = source_wb[sheet_name]
            new_sheet_name = f"room_{room}_{sheet_name}"
            # 确保工作表名不超过31个字符
            if len(new_sheet_name) > 31:
                new_sheet_name = new_sheet_name[:31]

            # 处理重复的工作表名
            counter = 1
            original_new_name = new_sheet_name
            while new_sheet_name in wb.sheetnames:
                new_sheet_name = f"{original_new_name}_{counter}"
                counter += 1

            target_ws = wb.create_sheet(title=new_sheet_name)

            # 1. 复制单元格数据
            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(row)

            # 2. 复制图片（关键新增逻辑）
            from openpyxl.drawing.image import Image
            import io
            for img in source_ws._images:
                # 获取图片在源工作表中的位置
                anchor = img.anchor

                # 正确获取图片数据
                img_bytes = img._data()  # 调用函数获取字节数据
                img_stream = io.BytesIO(img_bytes)  # 创建字节流

                # 创建新图片对象
                new_img = Image(img_stream)
                # 保持图片位置不变
                new_img.anchor = anchor
                # 添加到目标工作表
                target_ws.add_image(new_img)

            # 3. 复制列宽和行高
            for col in source_ws.column_dimensions:
                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
            for row in source_ws.row_dimensions:
                target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

            source_wb.close()

        # 保存到临时内存
        temp_file = io.BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)

        # 提供下载
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"batch_download_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"批量下载表单出错: {str(e)}")
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'})


# 注册蓝图
app.register_blueprint(admin_bp)

if __name__ == '__main__':
    init_db()
    app.run(debug=True)